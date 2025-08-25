from openpyxl import load_workbook

file_path = "my_excel_file.xlsx"
sheet_name = "ExistingSheet"
try:
    workbook = load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' is confirmed to exist in '{file_path}'.")
    else:
        print(f"Sheet '{sheet_name}' does not exist in '{file_path}'.")
except FileNotFoundError:
    print(f"File '{file_path}' does not exist.")
except Exception as e:
    print(f"Error checking sheet existence with openpyxl: {e}")


my_list = ["apple", "banana", "cherry"]
substring = "an"

found = any(substring in item for item in my_list)
print(found)
